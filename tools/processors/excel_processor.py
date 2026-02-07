"""
Excel processor for editing and exporting spreadsheets
"""
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors


def excel_to_json(file_bytes):
    """Convert Excel file to JSON format for editing"""
    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    ws = wb.active
    
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append([str(cell) if cell is not None else '' for cell in row])
    
    return data


def json_to_excel(data):
    """Convert JSON data to Excel file bytes with formatting"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Define styles
    header_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    header_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='D0D7DE'),
        right=Side(style='thin', color='D0D7DE'),
        top=Side(style='thin', color='D0D7DE'),
        bottom=Side(style='thin', color='D0D7DE')
    )
    alignment = Alignment(horizontal='left', vertical='center')
    
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = border
            cell.alignment = alignment
            cell.font = Font(name='Calibri', size=11)
            
            # Style first row as header
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()



def json_to_pdf(data, title="Spreadsheet Export"):
    """Convert JSON data to PDF with Excel-like formatting and proper text wrapping"""
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib.units import inch
    from reportlab.platypus import Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER
    
    output = BytesIO()
    doc = SimpleDocTemplate(
        output, 
        pagesize=landscape(A4),
        leftMargin=0.3*inch,
        rightMargin=0.3*inch,
        topMargin=0.3*inch,
        bottomMargin=0.3*inch
    )
    
    # Filter out completely empty rows
    filtered_data = []
    for row in data:
        if any(str(cell).strip() for cell in row):
            filtered_data.append(row)
    
    if not filtered_data:
        filtered_data = [['No data']]
    
    # Calculate column widths
    max_cols = max(len(row) for row in filtered_data)
    available_width = landscape(A4)[0] - 0.6*inch
    col_width = available_width / max_cols if max_cols > 0 else 1*inch
    
    # Ensure all rows have same columns
    for row in filtered_data:
        while len(row) < max_cols:
            row.append('')
    
    # Create paragraph styles for text wrapping
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontSize=8,
        alignment=TA_LEFT,
        fontName='Helvetica'
    )
    
    # Convert text to Paragraphs for wrapping
    table_data = []
    for row_idx, row in enumerate(filtered_data):
        new_row = []
        for cell in row:
            cell_text = str(cell) if cell else ''
            # Limit cell text length
            if len(cell_text) > 100:
                cell_text = cell_text[:97] + '...'
            
            if row_idx == 0:  # Header
                new_row.append(Paragraph(cell_text, header_style))
            else:
                new_row.append(Paragraph(cell_text, cell_style))
        table_data.append(new_row)
    
    # Create table
    table = Table(table_data, colWidths=[col_width] * max_cols, repeatRows=1)
    
    # Excel-like styling
    table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1f2937')),
        
        # Data cells
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1f2937')),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d0d7de')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        
        # Padding
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    
    doc.build([table])
    output.seek(0)
    return output.getvalue()



def create_empty_spreadsheet(rows=20, cols=10):
    """Create empty spreadsheet data"""
    return [['' for _ in range(cols)] for _ in range(rows)]
